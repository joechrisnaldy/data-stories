"""Compute every number quoted in Post 23 and write results.json.

Nothing in the draft may be a figure recalled from memory. If a number appears in the post, it
appears here first. Run:  python3 build_analysis.py
"""

import csv
import json
import os

import numpy as np
import openpyxl

from conditions import CLASS_ORDER, CONDITIONS, tclass

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")

WBINCOME = os.path.join(DATA, "ghe2021_daly_wbincome_new.xlsx")
BYCOUNTRY = {y: os.path.join(DATA, f"ghe2021_daly_bycountry_{y}.xlsx") for y in (2010, 2021)}
WHODEATHS = os.path.join(DATA, "ghe2021_deaths_global_new2.xlsx")
USA = "United States of America"


# ------------------------------------------------------------------ WHO GHE parsing

def income_sheet(sheet):
    """{cause name: DALYs} plus population, from the income-group workbook.

    Cause names, never codes. GHE code 970 is Epilepsy and 980 is Multiple sclerosis, the
    reverse of the obvious guess, so codes are not used anywhere in this file.
    """
    wb = openpyxl.load_workbook(WBINCOME, read_only=True, data_only=True)
    ws = wb[sheet]
    causes, pop = {}, None
    for r in ws.iter_rows(min_row=6, max_col=7, values_only=True):
        label = r[0]
        if label and str(label).startswith("Population"):
            pop = r[6] * 1000.0          # sheet is in thousands
        if r[0] is None:
            continue
        name = None
        for c in r[1:6]:
            if c not in (None, ""):
                name = str(c).strip()
        if name and isinstance(r[6], (int, float)):
            causes[name] = float(r[6])
    assert pop, f"{sheet}: population row not found"
    assert "All Causes" in causes, f"{sheet}: All Causes not found"
    return causes, pop


def country_sheet(year, country=USA):
    """{cause name: DALYs} plus population, for one country, from the by-country workbook.

    Wide layout: row 7 holds country names, row 8 the ISO-3 codes, and the value columns are in
    thousands of DALYs. Only the 'Persons' rows are read.
    """
    wb = openpyxl.load_workbook(BYCOUNTRY[year], read_only=True, data_only=True)
    ws = wb["All ages"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[6]
    col = next((j for j, v in enumerate(header) if v and str(v).strip() == country), None)
    assert col is not None, f"{year}: column for {country!r} not found"

    causes, pop = {}, None
    for r in rows[7:]:
        if r[0] != "Persons":
            continue
        name = None
        for c in r[2:7]:
            if c not in (None, ""):
                name = str(c).strip()
        val = r[col] if col < len(r) else None
        if not isinstance(val, (int, float)):
            continue
        if name and name.startswith("Population"):
            pop = float(val) * 1000.0
        elif name:
            causes[name] = float(val) * 1000.0
    assert pop, f"{year} {country}: population not found"
    assert "All Causes" in causes, f"{year} {country}: All Causes not found"
    return causes, pop


def owid(slug):
    with open(os.path.join(DATA, f"{slug}.csv"), encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def series(slug, entity, value_col=None, year_col="year"):
    rows = [r for r in owid(slug) if r.get("entity") == entity]
    assert rows, f"{slug}: entity {entity!r} not present"
    col = value_col or [c for c in rows[0] if c not in ("entity", "code", "year", "day")][0]
    out = {}
    for r in rows:
        if r[col] not in ("", None):
            out[int(str(r[year_col])[:4])] = float(r[col])
    return out


# ------------------------------------------------------------------ small stats helpers

def ols(X, y):
    X, y = np.asarray(X, float), np.asarray(y, float)
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    n, k = X.shape
    dof = n - k
    XtXi = np.linalg.pinv(X.T @ X)
    se = np.sqrt(np.diag(XtXi) * (resid @ resid) / dof)
    ss_tot = float(((y - y.mean()) ** 2).sum())
    r2 = 1.0 - float(resid @ resid) / ss_tot
    return dict(beta=[round(b, 4) for b in beta],
                se=[round(s, 4) for s in se],
                t=[round(b / s, 2) for b, s in zip(beta, se)],
                r2=round(r2, 4),
                adj_r2=round(1 - (1 - r2) * (n - 1) / dof, 4),
                n=n)


def median(xs):
    xs = sorted(xs)
    n = len(xs)
    return xs[n // 2] if n % 2 else (xs[n // 2 - 1] + xs[n // 2]) / 2


def spearman(a, b):
    def rank(v):
        order = sorted(range(len(v)), key=lambda i: v[i])
        r = [0.0] * len(v)
        i = 0
        while i < len(order):
            j = i
            while j + 1 < len(order) and v[order[j + 1]] == v[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r
    ra, rb = rank(a), rank(b)
    return float(np.corrcoef(ra, rb)[0, 1])


# ------------------------------------------------------------------ the condition table

def who_road_deaths():
    """WHO's OWN published road injury deaths, from the Global Health Estimates deaths tables.

    Added in fact-check round 6. The post's opening two numbers used to come from Our World in
    Data's aggregation, attributed in the prose to "the World Health Organization's Global Health
    Estimates". WHO publishes those estimates itself and says something different: 1,184,514 for
    2000 against OWID's 1,177,422. OWID's country values differ from WHO's published country file
    in 177 of 183 countries, so this is not a rounding difference, it is a different aggregation.
    A mirror is not the publisher, and round 1 of this fact-check caught the same shape with the
    SDG road-death rate.

    WHO publishes 2000, 2010, 2015, 2019, 2020 and 2021 in this file, not an annual series, so the
    year-by-year shape in the method notes stays on OWID's copy and says so.
    """
    wb = openpyxl.load_workbook(WHODEATHS, read_only=True, data_only=True)
    ws = wb["Summary"]
    for r in ws.iter_rows(min_row=9, max_col=10, values_only=True):
        if r[0] == 1530:                      # GHE cause code 1530, "Road injury"
            label = next(c for c in r[1:6] if isinstance(c, str) and "Road" in c)
            nums = [c for c in r[5:] if isinstance(c, (int, float))]
            assert label.strip() == "Road injury", label
            return {2021: nums[0] * 1000, 2000: nums[3] * 1000, "cause_label": label.strip()}
    raise AssertionError("GHE cause code 1530 (Road injury) not found in the deaths Summary sheet")


def burden_ranking(sheet="HI 2021", top=8):
    """Rank every GHE cause at the same depth, so the post never quotes a rank it cannot show.

    Added after fact-check round 1: the draft asserted that back and neck pain is the third
    largest cause in high-income countries, behind COVID-19 and ischaemic heart disease. The
    figure was computed in a throwaway shell command and appeared in no script, which breaks the
    rule that every number in a deliverable is reproduced by code in the repository.
    """
    wb = openpyxl.load_workbook(WBINCOME, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(min_row=9, max_col=7, values_only=True):
        if r[0] is None:
            continue
        depth = name = None
        for i, c in enumerate(r[1:6]):
            if c not in (None, ""):
                name, depth = str(c).strip(), i
        if name and isinstance(r[6], (int, float)):
            rows.append((name, depth, float(r[6])))
    target = next(x for x in rows if x[0] == "Back and neck pain")
    peers = sorted([x for x in rows if x[1] == target[1]], key=lambda x: -x[2])
    return dict(sheet=sheet, depth=target[1], n_causes_at_depth=len(peers),
                rank_of_back_and_neck_pain=[p[0] for p in peers].index("Back and neck pain") + 1,
                top=[dict(cause=p[0], dalys=p[2]) for p in peers[:top]])


def build_conditions():
    hi, hi_pop = income_sheet("HI 2021")
    gl, _ = income_sheet("Global 2021")
    us, us_pop = country_sheet(2021)

    ct = json.load(open(os.path.join(DATA, "ct_counts.json")))
    nih = json.load(open(os.path.join(DATA, "nih_spend.json")))
    spend = nih["totals"]

    rows = []
    for c in CONDITIONS:
        name = c["ghe_name"]
        assert name in hi, f"GHE high-income sheet has no cause {name!r}"
        assert name in gl, f"GHE global sheet has no cause {name!r}"
        assert name in us, f"GHE United States sheet has no cause {name!r}"
        assert name in ct, f"no ClinicalTrials.gov counts for {name!r}"

        terms = ct[name]
        best_term = max(terms, key=terms.get)
        trials = terms[best_term]
        assert trials > 0, f"{name}: zero trials, term list is wrong"

        primary = c["nih"]
        alts = {a: spend[a]["award_total_usd"] for a in c.get("nih_alt", []) if a in spend}
        usd = spend[primary]["award_total_usd"] if primary else None
        if primary:
            assert not spend[primary]["truncated"], f"{primary}: NIH paging truncated"

        rows.append(dict(
            cause=name, tclass=tclass(c), t1=c["t1"], t2=c["t2"], note=c["note"],
            hi_dalys=hi[name], global_dalys=gl[name], us_dalys=us[name],
            market_share_hi=hi[name] / gl[name],
            trials=trials, trials_term=best_term, trials_all_terms=terms,
            trials_per_m_hi=trials / (hi[name] / 1e6),
            trials_per_m_global=trials / (gl[name] / 1e6),
            nih_category=primary, nih_usd=usd,
            # Malaria's US burden rounds to zero, so dollars per US DALY is undefined for it.
            # That is a finding, not a nuisance: NIH funds it anyway. Recorded, not imputed.
            nih_usd_per_us_daly=(usd / us[name]) if (usd and us[name] > 0) else None,
            us_burden_is_zero=us[name] == 0,
            nih_alt_usd=alts,
        ))
    return rows, dict(hi_pop=hi_pop, us_pop=us_pop,
                      hi_all_causes=hi["All Causes"], us_all_causes=us["All Causes"],
                      global_all_causes=gl["All Causes"])


# ------------------------------------------------------------------ acts

USER_LABEL = {"ROADUSERTYPE_RS-DDC-4WHEELS": "Occupants of four-wheeled vehicles",
              "ROADUSERTYPE_RS-DDC-2OR3WHEELS": "Riders of two or three wheelers",
              "ROADUSERTYPE_RS-DDC-PEDESTRIANS": "Pedestrians",
              "ROADUSERTYPE_RS-DDC-CYCLISTS": "Cyclists",
              "ROADUSERTYPE_RS-DDC-OTHER": "Other or unspecified"}

# WHO's OWN published global distribution, Global status report on road safety 2023, pages 10, 15
# and 17, which state it three times identically. Country-reported data for 2021.
#
# READ THIS BEFORE CHANGING ANY OF IT. These figures have been wrong twice.
# Rounds 1 to 3 published a recomputation from country-level RS_246, labelled "car occupants", at
# 22.8 percent. Round 4 replaced it with 30 percent taken from WHO's LAUNCH NEWS RELEASE of
# 13 December 2023 rather than from the report. That release contradicts the report it announces:
# it assigns 30 percent to "car and other 4-wheeled light vehicle occupants" when 30 percent is the
# report's figure for MOTORCYCLISTS, and it prints a 3 percent micro-mobility share that the report
# says does not exist globally. Round 5 caught it by opening the report. The report is the source.
# A press release is not the report, and the IRIS full text answers a plain curl.
WHO_PUBLISHED_SPLIT = {"Riders of powered two and three wheelers": 30.0,
                       "Occupants of four-wheeled vehicles": 25.0,
                       "Pedestrians": 21.0,
                       "Cyclists": 5.0}
# WHO's own residual, published by WHO rather than derived here, with WHO's own composition.
WHO_PUBLISHED_RESIDUAL = 19.0
WHO_PUBLISHED_RESIDUAL_LABEL = "Buses, heavy goods, other and unknown"
WHO_PUBLISHED_RESIDUAL_NOTE = ('WHO: "Occupants of vehicles carrying more than 10 people, heavy '
                               'goods vehicles, \u201cother\u201d users and \u201cunknown\u201d user '
                               'types comprise the remaining 19% of deaths."')
# WHO states a WORD, not a number, for vulnerable road users: the report says pedestrians, cyclists
# and other vulnerable road users "account for half of all deaths". The 53 percent the post used to
# quote is the news release's, and it is built from the transposed shares. Do not reinstate it.
WHO_VULNERABLE_WORDING = "half of all deaths"
WHO_PUBLISHED_SOURCE = ("World Health Organization, Global status report on road safety 2023, "
                        "pages 10, 15 and 17; country-reported data for 2021")


def transport_act():
    """Road deaths from WHO directly, not from a mirror.

    REBUILT after fact-check round 1. The first version used Our World in Data's copy of the SDG
    road-death rate, which stops at 2019, and called 2019 "the most recent year the World Health
    Organization has published". WHO has published 2021. Worse, the rate fell about a sixth over
    the decade, so the original "road deaths barely moved" was beaten by WHO's own newer figure.

    What replaces it is stronger and comes from the same Global Health Estimates release as every
    disease-burden number in this post, so it is one consistent series rather than two vintages
    spliced: the COUNT of people killed is almost exactly the same in 2021 as in 2000. The rate
    per person fell because there are more people, not because fewer die.
    """
    ev = series("electric-car-sales-share", "World")
    # OWID's aggregation. Used ONLY for the year-by-year shape, because WHO publishes six years and
    # not an annual series. Every headline figure comes from WHO's own table, below.
    deaths = series("deaths-from-road-injuries", "World")
    who_deaths = who_road_deaths()

    rate_rows = json.load(open(os.path.join(DATA, "who_road_death_rate.json")))
    glob = [r for r in rate_rows if r["SpatialDim"] == "GLOBAL" and r.get("NumericValue")]
    assert len(glob) == 1, f"expected one GLOBAL row in RS_198, got {len(glob)}"
    rate_2021, rate_year = glob[0]["NumericValue"], glob[0]["TimeDim"]

    count_rows = json.load(open(os.path.join(DATA, "who_road_deaths_count.json")))
    gsrrs_total = sum(r["NumericValue"] for r in count_rows if r.get("NumericValue"))

    # The count barely moved and the per-person rate fell. The post says both, so both have to be
    # computed here rather than recalled: fact-check round 5 found the rate sentence resting on UN
    # population figures typed into a scratch script and reproducible from nothing in this
    # repository. World population comes from the SAME Global Health Estimates release as every
    # other figure in this post, whose Global sheets carry a population row in thousands.
    pop_2000 = income_sheet("Global 2000")[1]
    pop_2021 = income_sheet("Global 2021")[1]
    rate_2000 = who_deaths[2000] / pop_2000 * 1e5
    rate_2021_computed = who_deaths[2021] / pop_2021 * 1e5

    # Share of deaths by road user type.
    #
    # REBUILT AGAIN after fact-check round 4, and this is the important one. The post used to lead
    # with a figure computed here from country-level RS_246: 2016 shares weighted by each country's
    # 2021 death count. That gave car occupants 22.8 percent. WHO's own Global status report on
    # road safety 2023, which is the upstream source of RS_246, publishes 30 percent. The gap is
    # the completeness filter: requiring all five categories present drops the sample from 139
    # countries covering 66 percent of the world's road deaths to 82 covering 37 percent, and the
    # countries that survive are not a random third. Publishing a subsample recomputation in bold,
    # 7 points below the source's own published figure and in the direction that flatters the
    # argument, is the exact failure this post is about. WHO's published split is now primary; the
    # recomputation is kept, computed on BOTH filters, and reported as a check that disagrees.
    user_rows = json.load(open(os.path.join(DATA, "who_road_user_type.json")))
    by_country = {}
    for r in user_rows:
        if r["TimeDim"] == 2016 and r.get("NumericValue") is not None:
            by_country.setdefault(r["SpatialDim"], {})[r["Dim1"]] = r["NumericValue"]
    dcount = {r["SpatialDim"]: r["NumericValue"] for r in count_rows if r.get("NumericValue")}
    world_deaths = sum(dcount.values())

    def weighted_split(sel):
        """Death-weighted mean of each country's split, renormalised to its own reported total."""
        agg, cov = {k: 0.0 for k in USER_LABEL}, 0.0
        for c in sel:
            v, s = by_country[c], sum(by_country[c].values())
            if s <= 0:
                continue
            cov += dcount[c]
            for k, val in v.items():
                agg[k] += dcount[c] * val / s
        sh = {USER_LABEL[k]: 100 * v / cov for k, v in agg.items()}
        return sh, cov, len(sel)

    assert by_country, "RS_246 parsed to nothing: check the 2016 filter and the dimension codes"
    complete = [c for c, v in by_country.items()
                if c in dcount and len(v) == len(USER_LABEL) and abs(sum(v.values()) - 100) < 12]
    reporting = [c for c, v in by_country.items() if c in dcount and sum(v.values()) > 0]
    assert complete, "no country passes the complete-split filter"
    shares, covered, n_complete = weighted_split(complete)
    # NOT an assert on the recomputed shares summing to 100: weighted_split divides by the sum of
    # its own numerator, so that identity cannot fail and testing it is decoration. Round 6.
    # This one can fail, and would have caught round 4's actual error, which transposed two
    # categories while leaving the total at 100.
    assert WHO_PUBLISHED_SPLIT["Riders of powered two and three wheelers"] > \
        WHO_PUBLISHED_SPLIT["Occupants of four-wheeled vehicles"], \
        "WHO's report puts riders above four-wheel occupants; a transposition has crept back in"
    loose_shares, loose_covered, n_reporting = weighted_split(reporting)

    def vuln(sh):
        # Keyed off USER_LABEL, never off a literal. Round 5 renamed the four-wheel label and left
        # a dead "Car occupants" string here, so this counted four-wheel occupants as vulnerable
        # and returned 79 where it should return 56. Round 6 caught it. The two excluded keys are
        # derived from the label map so a rename can never silently change the meaning again.
        not_vulnerable = {USER_LABEL["ROADUSERTYPE_RS-DDC-4WHEELS"],
                          USER_LABEL["ROADUSERTYPE_RS-DDC-OTHER"]}
        return sum(v for k, v in sh.items() if k not in not_vulnerable)

    assert abs(vuln(shares) - (100 - shares[USER_LABEL["ROADUSERTYPE_RS-DDC-4WHEELS"]]
                               - shares[USER_LABEL["ROADUSERTYPE_RS-DDC-OTHER"]])) < 1e-9, \
        "vuln() is not excluding the categories it claims to"

    # No assertion on the computed vulnerable share. The old code asserted it exceeded 50 percent,
    # which passed only on the filtered sample and so validated the sample rather than testing it:
    # on the 139 reporting countries it is 49.2. An assertion that can only be satisfied by the
    # sample you chose is not a check.
    # Every honest variant of the recomputation, so the post cannot quote a flattering one.
    # WITHDRAWN in round 5: the "all 139 reporting countries" variant. Six of those countries
    # returned exactly ONE category, all of them four-wheel occupants, and renormalising a single
    # category to its own sum scales it to 100 percent. Those six contribute 34.8 percent of the
    # 31.3 percent that variant produced. Excluding them it returns 22.9, which is the complete
    # filter's answer. It was never a second estimate, it was the first estimate plus an artefact,
    # and round 4 published it as the upper end of a bracket around WHO's figure.
    no_singleton = [c for c in reporting if len(by_country[c]) > 1]
    all_five = [c for c in by_country if c in dcount and len(by_country[c]) == len(USER_LABEL)]
    variants = {}
    for key, sel in (("complete_filter", complete), ("all_five_no_tolerance", all_five),
                     ("reporting_excluding_single_category", no_singleton),
                     ("reporting_all_WITHDRAWN_artefact", reporting)):
        sh, cov, n = weighted_split(sel)
        variants[key] = dict(sh, **{"vulnerable": vuln(sh), "countries": n,
                                    "coverage_pct": 100 * cov / world_deaths})

    disagreement = dict(
        published=dict(WHO_PUBLISHED_SPLIT,
                       **{WHO_PUBLISHED_RESIDUAL_LABEL: WHO_PUBLISHED_RESIDUAL,
                          "vulnerable_road_users": WHO_VULNERABLE_WORDING}),
        published_source=WHO_PUBLISHED_SOURCE,
        variants=variants,
        filter_steps=dict(reporting_anything=len(reporting), all_five_categories=len(all_five),
                          and_summing_near_100=n_complete,
                          tolerance="the five shares must sum to within 12 points of 100"),
        note=("WHO's report puts occupants of four-wheeled vehicles at 25 percent. Recomputing "
              "from country-level RS_246, whose returns are 2013 and 2016, gives 22.8 percent on "
              "the 82 countries reporting a complete five-way split that sums near 100, 26.2 on "
              "all 101 reporting five categories, and 22.9 on all countries reporting more than "
              "one. The recomputation is about two points below WHO and the discrepancy is "
              "sample coverage, not a different answer. The 31.3 percent an earlier draft quoted "
              "as an upper bound is withdrawn: it is an artefact of renormalising six countries "
              "that reported a single category."))

    y0, y1 = min(deaths), max(deaths)
    return dict(
        source=("WHO Global Health Observatory indicators RS_196, RS_198 and RS_246 fetched "
                "directly; WHO Global Health Estimates 2021 for the death count series; "
                "International Energy Agency via Our World in Data for electric car sales"),
        # PRIMARY, WHO's own published deaths table. These are the two numbers the post opens on.
        deaths_first=who_deaths[2000], deaths_last=who_deaths[2021],
        deaths_first_year=2000, deaths_last_year=2021,
        deaths_pct_change=100 * (who_deaths[2021] / who_deaths[2000] - 1),
        deaths_source="WHO Global Health Estimates 2021, deaths tables, cause code 1530 Road injury",
        # SECONDARY, Our World in Data's annual aggregation of the same estimates. It differs from
        # WHO's own table by roughly 7,000 to 9,000 deaths a year, so it is kept for the shape of
        # the series and never quoted as a level.
        road_deaths_by_year={str(k): v for k, v in sorted(deaths.items())},
        owid_series_note=("Our World in Data's aggregation, used for the year-by-year shape only. "
                          f"Its 2021 total is {deaths[y1]:,.0f} against WHO's own "
                          f"{who_deaths[2021]:,.0f}."),
        owid_deaths_first=deaths[y0], owid_deaths_last=deaths[y1],
        deaths_2010=deaths[2010], deaths_2021=deaths[2021],
        deaths_pct_change_2010_2021=100 * (deaths[2021] / deaths[2010] - 1),
        who_rate_per_100k_2021=rate_2021, who_rate_year=rate_year,
        who_gsrrs_country_sum=gsrrs_total,
        world_pop_2000=pop_2000, world_pop_2021=pop_2021,
        world_pop_growth_pct=100 * (pop_2021 / pop_2000 - 1),
        rate_per_100k_2000=rate_2000, rate_per_100k_2021_computed=rate_2021_computed,
        rate_change_pct_2000_2021=100 * (rate_2021_computed / rate_2000 - 1),
        # PRIMARY: WHO's own published global split. Everything the post says about who dies on
        # the roads is quoted from here.
        user_shares_pct=dict(WHO_PUBLISHED_SPLIT,
                             **{WHO_PUBLISHED_RESIDUAL_LABEL: WHO_PUBLISHED_RESIDUAL}),
        user_shares_published_only=dict(WHO_PUBLISHED_SPLIT),
        user_shares_residual_pct=WHO_PUBLISHED_RESIDUAL,
        user_shares_residual_note=WHO_PUBLISHED_RESIDUAL_NOTE,
        vulnerable_wording=WHO_VULNERABLE_WORDING,
        not_in_four_wheeler_pct=round(
            100 - WHO_PUBLISHED_SPLIT["Occupants of four-wheeled vehicles"], 1),
        user_shares_source=WHO_PUBLISHED_SOURCE,
        user_shares_residual_label=WHO_PUBLISHED_RESIDUAL_LABEL,
        user_share_year=2021,
        # SECONDARY: the country-level recomputation, on both filters, and what it disagrees about.
        user_split_disagreement=disagreement,
        user_share_countries=n_complete, user_share_deaths_covered=covered,
        user_share_coverage_pct=100 * covered / world_deaths,
        ev_share_by_year={str(k): v for k, v in sorted(ev.items())},
        ev_first_year=min(ev), ev_last_year=max(ev),
        ev_share_first=ev[min(ev)], ev_share_last=ev[max(ev)],
        ev_share_2010=ev.get(2010), ev_share_2021=ev.get(2021),
    )


def ai_act():
    priv = series("private-investment-in-artificial-intelligence", "World")
    corp = series("corporate-investment-in-artificial-intelligence-by-type", "Total")
    y0, y1 = min(priv), max(priv)
    rows = owid("computation-used-to-train-notable-artificial-intelligence-systems")
    flops = [(r["entity"], r["day"], float(r["training_computation_petaflop"]))
             for r in rows if r.get("training_computation_petaflop")]
    flops.sort(key=lambda x: x[1])
    top = max(flops, key=lambda x: x[2])
    early = [f for f in flops if f[1][:4] <= "2012"]
    return dict(
        source="Stanford AI Index / Quid and Epoch AI, both via Our World in Data",
        private_investment_usd={str(k): v for k, v in sorted(priv.items())},
        first_year=y0, last_year=y1,
        private_first=priv[y0], private_last=priv[y1],
        private_multiple=priv[y1] / priv[y0],
        private_pct_change=100 * (priv[y1] / priv[y0] - 1),
        corporate_total_last=corp[max(corp)], corporate_last_year=max(corp),
        n_models=len(flops),
        largest_model=top[0], largest_model_date=top[1], largest_model_petaflop=top[2],
        earliest_model=early[0][0] if early else None,
        earliest_model_date=early[0][1] if early else None,
        earliest_model_petaflop=early[0][2] if early else None,
        compute_multiple=(top[2] / early[0][2]) if early else None,
        benefit_series_available=False,
        # The named figures here are external, published, and quoted in the post with a reference.
        # They are recorded in this file so that the number in the essay and the number in the
        # chart come from one place, per the house rule that no figure is quoted from memory.
        # From the STUDY, not from the AI Index's restatement of it. The Index prints $112bn for
        # 2025 because it plots a 95 million user base; the authors use 98.78 million and get
        # $116.2bn. Round 5 corrected the post and not this file, so the essay briefly quoted a
        # number the repository could not produce. Round 6 caught that.
        consumer_surplus_usd={"2025-07": 116.2e9, "2026-03": 172.3e9},
        consumer_surplus_pct_change=48.3,
        consumer_surplus_source=("Brynjolfsson, Collis, Eggers, Kazinnik and Nguyen, What is "
                                 "Generative AI Worth?, 13 April 2026: willingness-to-accept "
                                 "elicited from representative samples of US adults in two waves, "
                                 "July 2025 and March 2026; aggregate consumer surplus "
                                 "$116 billion to $172 billion"),
        benefit_note=("CORRECTED TWICE. Round 1 killed 'no measure of AI's effect on people "
                      "exists': randomised task-level studies exist. Round 4 killed the "
                      "replacement, 'no population-level welfare series exists', which round 3 "
                      "produced by dropping the word AGREED: the Stanford AI Index, the same "
                      "publication this post's investment line comes from, carries a "
                      "population-level US consumer-surplus estimate with two dated observations. "
                      "What survives, and all the post now claims, is that there is no AGREED "
                      "year-by-year welfare series to run alongside the money: what exists is one "
                      "research group's stated-preference estimate, US only, two observations a "
                      "year apart, first published in 2026 against a money line starting in 2013."),
    )


def synthesis(transport, ai):
    """Chart 4 rows. Windows differ by domain and every row carries its own window."""
    hi10, pop10 = income_sheet("HI 2010")
    hi19, pop19 = income_sheet("HI 2019")
    rate10 = hi10["All Causes"] / pop10 * 1e5
    rate19 = hi19["All Causes"] / pop19 * 1e5
    by_year = json.load(open(os.path.join(DATA, "ct_by_year.json")))

    return [
        dict(domain="Transport", window="2010 to 2021",
             input_label="Electric share of new car sales",
             input_first=transport["ev_share_2010"], input_last=transport["ev_share_2021"],
             input_pct_change=100 * (transport["ev_share_2021"] / transport["ev_share_2010"] - 1),
             outcome_label="People killed on the roads",
             outcome_first=transport["deaths_2010"], outcome_last=transport["deaths_2021"],
             outcome_pct_change=transport["deaths_pct_change_2010_2021"]),
        dict(domain="Medicine", window="2010 to 2019",
             input_label="Clinical trials first registered that year",
             input_first=by_year["2010"], input_last=by_year["2019"],
             input_pct_change=100 * (by_year["2019"] / by_year["2010"] - 1),
             outcome_label="Healthy years lost per 100,000, high-income countries",
             outcome_first=rate10, outcome_last=rate19,
             outcome_pct_change=100 * (rate19 / rate10 - 1)),
        dict(domain="Artificial intelligence", window=f"{ai['first_year']} to {ai['last_year']}",
             input_label="Private investment",
             input_first=ai["private_first"], input_last=ai["private_last"],
             input_pct_change=ai["private_pct_change"],
             outcome_label="No agreed year-by-year welfare series exists",
             outcome_first=None, outcome_last=None, outcome_pct_change=None),
    ]


# ------------------------------------------------------------------ models

def money_rows(rows):
    """Rows for the money analysis, with shared RCDC categories MERGED rather than deduplicated.

    Road injury and falls are both tagged to "Physical Injury - Accidents and Adverse Effects",
    so that category's dollars pay for both. Keeping one condition and discarding the other
    would credit the survivor with the pair's whole budget against half the harm, which roughly
    doubles its dollars per healthy year lost. The two are summed into one row instead.

    Also dropped here: leukaemia, which has no general RCDC category, and malaria, whose US
    burden rounds to zero so dollars per US healthy year lost is undefined.
    """
    merged = {}
    for r in sorted([x for x in rows if x["nih_usd"]], key=lambda x: -x["us_dalys"]):
        key = r["nih_category"]
        if key in merged:
            m = merged[key]
            m["cause"] = f"{m['cause']} and {r['cause'][0].lower() + r['cause'][1:]}"
            for f in ("hi_dalys", "global_dalys", "us_dalys", "trials"):
                m[f] += r[f]
            m["merged_from"].append(r["cause"])
            assert m["t1"] == r["t1"], f"{key}: merged conditions disagree on the target axis"
        else:
            merged[key] = dict(r, merged_from=[r["cause"]])
    out = []
    for m in merged.values():
        if m["us_dalys"] <= 0:
            continue
        m["market_share_hi"] = m["hi_dalys"] / m["global_dalys"]
        m["nih_usd_per_us_daly"] = m["nih_usd"] / m["us_dalys"]
        m["trials_per_m_hi"] = m["trials"] / (m["hi_dalys"] / 1e6)
        out.append(m)
    return sorted(out, key=lambda r: -r["hi_dalys"])


def models(rows):
    lg = np.log10
    y = lg([r["trials"] for r in rows])
    burden = lg([r["hi_dalys"] for r in rows])
    market = np.array([r["market_share_hi"] for r in rows])
    partly = np.array([1.0 if r["tclass"] == "partly" else 0.0 for r in rows])
    intract = np.array([1.0 if r["tclass"] == "intractable" else 0.0 for r in rows])
    t1 = np.array([0.0 if r["t1"] else 1.0 for r in rows])   # 1 = NO validated target
    one = np.ones(len(rows))

    out = {
        "m1_burden_only": dict(terms=["const", "log10 HI DALYs"],
                               **ols(np.column_stack([one, burden]), y)),
        "m2_burden_plus_market": dict(terms=["const", "log10 HI DALYs", "high-income share"],
                                      **ols(np.column_stack([one, burden, market]), y)),
        "m3_burden_plus_tractability": dict(
            terms=["const", "log10 HI DALYs", "partly (vs tractable)", "intractable (vs tractable)"],
            **ols(np.column_stack([one, burden, partly, intract]), y)),
        "m4_all": dict(
            terms=["const", "log10 HI DALYs", "high-income share", "partly", "intractable"],
            **ols(np.column_stack([one, burden, market, partly, intract]), y)),
        "m5_no_target_only": dict(terms=["const", "log10 HI DALYs", "no validated target"],
                                  **ols(np.column_stack([one, burden, t1]), y)),
    }

    # Sensitivity to the coding judgement a reader is most likely to dispute. Road injury and
    # falls are coded t1=False on "no biological target"; a reader who counts a crash as an
    # identified causal agent, the rule's own second limb, would code them True. The
    # pre-registration forbids recoding after seeing a residual, so the coding does not change.
    # This is what it would cost. Added in round 6, which found conditions.py and the design
    # document both asserting this number was already published when it was computed nowhere.
    out["recode_injuries_as_target"] = dict(
        terms=["const", "log10 HI DALYs", "no validated target, injuries recoded as having one"],
        **ols(np.column_stack([one, burden,
                               np.array([0.0 if (r["t1"] or r["cause"] in ("Road injury", "Falls"))
                                         else 1.0 for r in rows])]), y))

    # Sensitivity to the one condition whose registry mapping failed. Road injury is a CAUSE of
    # injury, and ClinicalTrials.gov has no condition corresponding to it: every pre-registered
    # term was crash vocabulary, which undercounts, while injury-pathology terms overcount
    # because most fractures and head injuries are not road-related. The point stays in, because
    # dropping an inconvenient observation after seeing its residual is the move this repository
    # refuses. The result of dropping it is reported instead.
    keep = [r for r in rows if r["cause"] != "Road injury"]
    for basis, tag in (("hi_dalys", "hi"), ("global_dalys", "global")):
        b = lg([r[basis] for r in keep])
        nt = np.array([0.0 if r["t1"] else 1.0 for r in keep])
        out[f"drop_road_injury_{tag}"] = dict(
            terms=["const", f"log10 {basis}", "no validated target"],
            **ols(np.column_stack([np.ones(len(keep)), b, nt]), lg([r["trials"] for r in keep])))

    # Same three models on the GLOBAL burden basis. Trials are a global count, so pairing them
    # with global burden is the like-for-like test; the high-income basis is the one that makes
    # the market question answerable. Both are reported and the post states which it quotes.
    bg = lg([r["global_dalys"] for r in rows])
    out["global_m1_burden_only"] = dict(terms=["const", "log10 global DALYs"],
                                        **ols(np.column_stack([one, bg]), y))
    out["global_m3_plus_tractability"] = dict(
        terms=["const", "log10 global DALYs", "partly", "intractable"],
        **ols(np.column_stack([one, bg, partly, intract]), y))
    out["global_m5_no_target_only"] = dict(
        terms=["const", "log10 global DALYs", "no validated target"],
        **ols(np.column_stack([one, bg, t1]), y))

    dedup = money_rows(rows)

    def money_models(sub, tag):
        yn = lg([r["nih_usd"] for r in sub])
        bn = lg([r["us_dalys"] for r in sub])
        mk = np.array([r["market_share_hi"] for r in sub])
        pn = np.array([1.0 if r["tclass"] == "partly" else 0.0 for r in sub])
        inn = np.array([1.0 if r["tclass"] == "intractable" else 0.0 for r in sub])
        on = np.ones(len(sub))
        out[f"{tag}_m1_burden_only"] = dict(terms=["const", "log10 US DALYs"],
                                            **ols(np.column_stack([on, bn]), yn))
        out[f"{tag}_m2_plus_market"] = dict(terms=["const", "log10 US DALYs", "high-income share"],
                                            **ols(np.column_stack([on, bn, mk]), yn))
        out[f"{tag}_m3_plus_tractability"] = dict(
            terms=["const", "log10 US DALYs", "partly", "intractable"],
            **ols(np.column_stack([on, bn, pn, inn]), yn))
        # The binary that works on trials, run on dollars. Round 1 refuted the draft's claim that
        # the money "pattern is the same but softer": it is not attenuated, it is absent.
        ntn = np.array([0.0 if r["t1"] else 1.0 for r in sub])
        out[f"{tag}_m4_no_target"] = dict(
            terms=["const", "log10 US DALYs", "no validated target"],
            **ols(np.column_stack([on, bn, ntn]), yn))
        out[f"{tag}_conditions_used"] = [r["cause"] for r in sub]

    money_models(dedup, "money")

    # NIH funds HIV, tuberculosis and malaria against GLOBAL burden, not American burden, which
    # is why they land at thousands of dollars per US DALY. Leaving them in makes the dollar test
    # a test of where disease is rather than of what research money follows. The restricted set
    # is reported alongside the full one, never instead of it.
    GLOBAL_HEALTH = {"HIV/AIDS", "Tuberculosis", "Malaria"}
    money_models([r for r in dedup if r["cause"] not in GLOBAL_HEALTH], "money_domestic")
    out["money_domestic_excluded"] = sorted(GLOBAL_HEALTH)
    return out


# ------------------------------------------------------------------ main

def main():
    rows, totals = build_conditions()

    # Dollar medians MUST come from the merged money rows, not from `rows`. Road injury and Falls
    # share one RCDC category, so in the unmerged list each carries the whole $887m over its own
    # burden alone and both are roughly doubled. Fact-check round 5 found the class dollar medians
    # computed that way while the post's method note promised the merged basis "so that neither is
    # credited with the other's money". It put the partly class at $288 instead of $160 and made
    # the top two classes look 4 percent apart when they are 47.
    mrows = {r["cause"]: r for r in money_rows(rows)}
    merged_of = {}
    for r in money_rows(rows):
        for c in r.get("merged_from", [r["cause"]]):
            merged_of[c] = r

    by_class = {}
    for cls in CLASS_ORDER:
        sub = [r for r in rows if r["tclass"] == cls]
        seen, money = set(), []
        for r in sub:
            m = merged_of.get(r["cause"])
            if m and m["cause"] not in seen and m["nih_usd_per_us_daly"]:
                seen.add(m["cause"])
                money.append(m["nih_usd_per_us_daly"])
        by_class[cls] = dict(
            n=len(sub), causes=[r["cause"] for r in sub],
            median_trials_per_m_hi=median([r["trials_per_m_hi"] for r in sub]),
            median_trials_per_m_global=median([r["trials_per_m_global"] for r in sub]),
            median_nih_usd_per_us_daly=median(money) if money else None,
            median_hi_dalys=median([r["hi_dalys"] for r in sub]),
        )

    by_target = {}
    for has in (True, False):
        sub = [r for r in rows if r["t1"] is has]
        seen, money = set(), []
        for r in sub:
            m = merged_of.get(r["cause"])
            if m and m["cause"] not in seen and m["nih_usd_per_us_daly"]:
                seen.add(m["cause"])
                money.append(m["nih_usd_per_us_daly"])
        by_target["validated_target" if has else "no_validated_target"] = dict(
            n=len(sub), causes=[r["cause"] for r in sub],
            median_trials_per_m_hi=median([r["trials_per_m_hi"] for r in sub]),
            median_trials_per_m_global=median([r["trials_per_m_global"] for r in sub]),
            median_nih_usd_per_us_daly=median(money) if money else None)

    look = {r["cause"]: r for r in rows}
    bp, ms, ra = look["Back and neck pain"], look["Multiple sclerosis"], look["Rheumatoid arthritis"]
    mal = look["Malaria"]

    headline = dict(
        back_pain_vs_ms=dict(
            burden_multiple=bp["hi_dalys"] / ms["hi_dalys"],
            trials_multiple=bp["trials"] / ms["trials"],
            money_multiple=bp["nih_usd"] / ms["nih_usd"],
            money_per_daly_multiple=ms["nih_usd_per_us_daly"] / bp["nih_usd_per_us_daly"],
            back_pain=dict(hi_dalys=bp["hi_dalys"], trials=bp["trials"], term=bp["trials_term"],
                           nih_usd=bp["nih_usd"], nih_category=bp["nih_category"],
                           nih_usd_per_us_daly=bp["nih_usd_per_us_daly"],
                           broader_category_usd=bp["nih_alt_usd"]),
            multiple_sclerosis=dict(hi_dalys=ms["hi_dalys"], trials=ms["trials"],
                                    nih_usd=ms["nih_usd"],
                                    nih_usd_per_us_daly=ms["nih_usd_per_us_daly"])),
        back_pain_vs_ra=dict(burden_multiple=bp["hi_dalys"] / ra["hi_dalys"],
                             trials_multiple=bp["trials"] / ra["trials"]),
        malaria=dict(hi_dalys=mal["hi_dalys"], global_dalys=mal["global_dalys"],
                     market_share_hi=mal["market_share_hi"], trials=mal["trials"],
                     nih_usd=mal["nih_usd"]),
        burden_rank_hi=[r["cause"] for r in sorted(rows, key=lambda r: -r["hi_dalys"])][:6],
        effort_rank_per_burden=[r["cause"] for r in sorted(rows, key=lambda r: -r["trials_per_m_hi"])][:6],
        lowest_effort_per_burden=[r["cause"] for r in sorted(rows, key=lambda r: r["trials_per_m_hi"])][:6],
        spearman_burden_vs_trials=spearman([r["hi_dalys"] for r in rows],
                                           [r["trials"] for r in rows]),
        spearman_burden_vs_trials_per_burden=spearman([r["hi_dalys"] for r in rows],
                                                      [r["trials_per_m_hi"] for r in rows]),
    )

    transport = transport_act()
    ai = ai_act()
    syn = synthesis(transport, ai)
    mods = models(rows)

    # ---- assertions catch a BROKEN BUILD, never a disagreeable finding ----
    # A failed hypothesis is a result to report, not an error to crash on. Only structural
    # breakage raises here; the pre-registered predictions are scored below and recorded
    # whether they passed or failed.
    #
    # Fact-check round 4 found this rule broken four lines below the comment stating it: three
    # assertions encoded claims the POST makes rather than structural facts about the build, so a
    # world in which back pain slipped to fourth, or road deaths finally moved, would have crashed
    # the build instead of changing the essay. Those are warnings now.
    assert len(rows) == len(CONDITIONS) == 34, len(rows)
    assert all(r["trials"] > 0 for r in rows)
    assert all(r["hi_dalys"] > 0 and r["global_dalys"] > 0 for r in rows)
    assert sum(1 for r in rows if r["nih_usd"] is None) == 1, \
        "exactly one condition (Leukaemia) should lack an NIH category"

    rank = burden_ranking()
    # Round 4 turned three assertions into printed warnings, correctly, because they encoded essay
    # claims rather than structural facts. Round 5 found the trade incomplete: a print in a build
    # nobody watches is not a check either. They are scored into results.json beside the
    # falsification conditions, which is what the scorecard exists for.
    post_claims = dict(
        back_pain_ranks_third=bool(rank["rank_of_back_and_neck_pain"] == 3),
        road_death_count_barely_moves=bool(abs(transport["deaths_pct_change"]) < 5),
        ai_investment_rises=bool(ai["private_multiple"] > 1),
    )
    for key, ok in post_claims.items():
        if not ok:
            print("WARNING, a claim in the post no longer holds:", key)

    med = {c: by_class[c]["median_trials_per_m_hi"] for c in CLASS_ORDER}
    scorecard = dict(
        post_claims_still_true=post_claims,
        full_ordering_predicted="tractable > partly > intractable",
        full_ordering_holds=med["tractable"] > med["partly"] > med["intractable"],
        tractable_above_intractable=med["tractable"] > med["intractable"],
        medians_trials_per_m_hi=med,
        target_binary_holds=(by_target["validated_target"]["median_trials_per_m_hi"] >
                             by_target["no_validated_target"]["median_trials_per_m_hi"]),
        market_beats_tractability_on_trials=(
            mods["m2_burden_plus_market"]["adj_r2"] > mods["m3_burden_plus_tractability"]["adj_r2"]),
        tractability_adds_over_burden_alone=(
            mods["m3_burden_plus_tractability"]["adj_r2"] > mods["m1_burden_only"]["adj_r2"]),
        market_adds_over_burden_alone=(
            mods["m2_burden_plus_market"]["adj_r2"] > mods["m1_burden_only"]["adj_r2"]),
        # bool(): round() on a numpy scalar returns np.float64, so the comparison yields np.bool_,
        # which is not an instance of bool and was therefore skipped by the scorecard printer and
        # written to results.json as 0.0. A pre-registered check silently vanishing from the
        # scorecard is exactly the failure round 2 found. Caught in round 4.
        burden_explains_nih_dollars=bool(abs(mods["money_m1_burden_only"]["t"][1]) > 2),

        # The three falsification conditions from design doc section 3.4, each scored explicitly.
        # Round 2 of the fact-check found that only the first was ever computed while the README
        # asserted all three had been checked. Silence is not a pass.
        fc1_market_beats_tractability=(
            mods["m2_burden_plus_market"]["adj_r2"] > mods["m3_burden_plus_tractability"]["adj_r2"]),
        fc2_gap_vanishes_under_generous_terms=(
            by_target["validated_target"]["median_trials_per_m_hi"]
            <= by_target["no_validated_target"]["median_trials_per_m_hi"]),
        fc2_caveat=("Terms are already maximised for every condition in the build, so this is a "
                    "live test rather than a hypothetical. It does not fire. But the maximise "
                    "rule provably failed for road injury, whose candidate terms were exhaustive "
                    "within crash vocabulary while the registry indexes trauma by pathology."),
        fc3_ordering_reverses_between_trials_and_dollars=(
            [c for c in sorted(CLASS_ORDER, key=lambda c: -by_class[c]["median_trials_per_m_hi"])]
            != [c for c in sorted(CLASS_ORDER,
                                  key=lambda c: -(by_class[c]["median_nih_usd_per_us_daly"] or 0))]),
        fc3_detail=dict(
            order_by_trials=sorted(CLASS_ORDER,
                                   key=lambda c: -by_class[c]["median_trials_per_m_hi"]),
            order_by_dollars=sorted(CLASS_ORDER,
                                    key=lambda c: -(by_class[c]["median_nih_usd_per_us_daly"] or 0)),
            note=("PARTIAL FIRE. The top class is the same on both measures; the middle and "
                  "bottom classes swap. Reported, not rounded away. Round 4 added the other half "
                  "of this, which cuts against the post's own narrative and was missing: the "
                  "three-step gradient the post says it abandoned is the ordering the DOLLARS "
                  "produce. It holds there and fails on trials. The post abandons it because it "
                  "fails on trials, which is the measure the argument rests on, and because the "
                  "top two dollar medians are 47 percent apart and the bottom two 15, so on "
                  "dollars the whole gradient is one step. Round 5 corrected this: it read "
                  "'4 percent apart, which is nothing', which was true only on the unmerged "
                  "basis where road injury and falls each carried the whole shared category.")),
        notes=("The three-class ordering FAILS: conditions coded intractable sit above those "
               "coded partly, because psychiatry runs large trial programmes with neither a "
               "validated target nor an objective endpoint. The binary target test survives and "
               "the market proxy adds nothing. The post reports all three."),
    )

    out = dict(
        meta=dict(
            post="Post 23: Greed You Can Regulate. Difficulty You Have to Pay For.",
            burden_source="WHO Global Health Estimates 2021, July 2024 release",
            effort_source="ClinicalTrials.gov API v2, study counts by condition",
            money_source=f"NIH RePORTER API v2, FY{json.load(open(os.path.join(DATA,'nih_spend.json')))['fiscal_year']} award totals by RCDC category",
            n_conditions=len(rows),
            caution=("RCDC categories overlap by construction and do not partition the NIH "
                     "budget. Trial counts are a registry-wide stock, not an annual flow. "
                     "Trials are global; DALYs are reported on high-income, global and US bases "
                     "and each figure states which."),
        ),
        totals=totals,
        conditions=rows,
        money_rows=money_rows(rows),
        by_class=by_class,
        by_target=by_target,
        burden_ranking=rank,
        scorecard=scorecard,
        headline=headline,
        models=mods,
        transport=transport,
        ai=ai,
        synthesis=syn,
    )
    with open(os.path.join(HERE, "results.json"), "w") as fh:
        json.dump(out, fh, indent=1, default=float)

    # ---- console summary ----
    print(f"{'condition':38s} {'class':12s} {'HI DALYs':>12s} {'trials':>7s} "
          f"{'tr/M':>8s} {'$/US DALY':>10s}")
    print("-" * 92)
    for r in sorted(rows, key=lambda r: -r["trials_per_m_hi"]):
        m = f"{r['nih_usd_per_us_daly']:.1f}" if r["nih_usd_per_us_daly"] else "n/a"
        print(f"{r['cause'][:38]:38s} {r['tclass']:12s} {r['hi_dalys']:>12,.0f} "
              f"{r['trials']:>7,} {r['trials_per_m_hi']:>8.0f} {m:>10s}")
    print("\nmedian by pre-registered class")
    for cls in CLASS_ORDER:
        b = by_class[cls]
        mm = f"${b['median_nih_usd_per_us_daly']:.1f}" if b["median_nih_usd_per_us_daly"] else "n/a"
        print(f"  {cls:12s} n={b['n']:<3d} trials/M HI = {b['median_trials_per_m_hi']:>8.0f}   "
              f"NIH $/US DALY = {mm}")
    print("\nmedian by validated-target binary")
    for k, b in by_target.items():
        mm = f"${b['median_nih_usd_per_us_daly']:.1f}" if b["median_nih_usd_per_us_daly"] else "n/a"
        print(f"  {k:22s} n={b['n']:<3d} trials/M HI = {b['median_trials_per_m_hi']:>8.0f}   "
              f"NIH $/US DALY = {mm}")

    def show(title, keys):
        print(f"\n{title}")
        for k in keys:
            ts = ", ".join(f"{t:+.2f}" for t in mods[k]["t"][1:])
            print(f"  {k:34s} n={mods[k]['n']:<3d} adjR2={mods[k]['adj_r2']:>7.3f}  t[{ts}]")

    show("dependent = log10 trials, high-income burden",
         ["m1_burden_only", "m2_burden_plus_market", "m3_burden_plus_tractability",
          "m4_all", "m5_no_target_only"])
    show("dependent = log10 trials, global burden",
         ["global_m1_burden_only", "global_m3_plus_tractability", "global_m5_no_target_only"])
    show("dependent = log10 NIH dollars, all conditions",
         ["money_m1_burden_only", "money_m2_plus_market", "money_m3_plus_tractability"])
    show(f"dependent = log10 NIH dollars, excluding {mods['money_domestic_excluded']}",
         ["money_domestic_m1_burden_only", "money_domestic_m2_plus_market",
          "money_domestic_m3_plus_tractability"])

    print("\nPRE-REGISTERED SCORECARD")
    for k, v in scorecard.items():
        # Only booleans are scored. A non-boolean entry is explanatory text, and printing PASS
        # beside it would manufacture a green tick out of a caveat.
        if not isinstance(v, bool):
            continue
        label = "FIRED" if k.startswith("fc") else ("PASS" if v else "FAIL")
        if k.startswith("fc"):
            label = "FIRED" if v else "did not fire"
        print(f"  {label:>12s}  {k}")
    print(f"  {'':>12s}  fc3 order by trials  = {scorecard['fc3_detail']['order_by_trials']}")
    print(f"  {'':>12s}  fc3 order by dollars = {scorecard['fc3_detail']['order_by_dollars']}")
    print(f"\n  {scorecard['notes']}")
    print(f"\nSpearman burden vs trials: {headline['spearman_burden_vs_trials']:.3f}")
    print("results.json written")


if __name__ == "__main__":
    main()
